import streamlit as st
import pandas as pd
import re
import io
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formataddr
from email import encoders
from openpyxl.styles import Alignment, PatternFill, Border, Side

# --- EMAIL CONFIGURATION ---
SENDER_EMAIL = "atharvaujoshi@gmail.com"
SENDER_NAME = "Spydarr Package Reporter" 
APP_PASSWORD = "nybl zsnx zvdw edqr"  

def send_email(recipient_email, excel_data, filename):
    try:
        recipient_name = recipient_email.split('@')[0].replace('.', ' ').title()
        msg = MIMEMultipart()
        msg['From'] = formataddr((SENDER_NAME, SENDER_EMAIL))
        msg['To'] = recipient_email
        msg['Subject'] = "Property Package Report"
        body = f"Dear {recipient_name},\n\nPlease find the attached professional Market Report with Package calculations.\n\nRegards,\nAtharva Joshi"
        msg.attach(MIMEText(body, 'plain'))
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(excel_data)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f"attachment; filename={filename}")
        msg.attach(part)
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(SENDER_EMAIL, APP_PASSWORD)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        st.error(f"Error sending email: {e}")
        return False

def extract_lower_carpet(value):
    if pd.isna(value): return 0
    numbers = re.findall(r"[-+]?\d*\.\d+|\d+", str(value))
    return float(numbers[0]) if numbers else 0

# --- UI SETUP ---
st.set_page_config(page_title="Package Calculator", layout="wide")
st.title("🏙️ Package Calculation & Professional Reporter")

uploaded_file = st.file_uploader("Upload your Excel file", type=['xlsx'])

if uploaded_file:
    try:
        xls = pd.ExcelFile(uploaded_file)
        target_sheet = next((s for s in xls.sheet_names if s.lower() == 'report'), None)
        
        if not target_sheet:
            st.error("Could not find a sheet named 'Report' or 'report'.")
        else:
            df = pd.read_excel(xls, sheet_name=target_sheet)
            
            # --- CALCULATIONS ---
            carpet_col = next((c for c in df.columns if "carpet area(sq.ft)" in c.lower()), None)
            apr_col = next((c for c in df.columns if "average of apr" in c.lower()), None)
            count_col = next((c for c in df.columns if "count of property" in c.lower()), None)

            if carpet_col and apr_col:
                lower_carpet = df[carpet_col].apply(extract_lower_carpet)
                df['Package'] = (lower_carpet * 1.568 * df[apr_col]).round(0) # 1.4 * 1.12 = 1.568
                
                if count_col:
                    cols = df.columns.tolist()
                    cols.remove('Package')
                    count_idx = cols.index(count_col)
                    cols.insert(count_idx, 'Package')
                    df = df[cols]

                # --- APPLY STYLING WITH OPENPYXL ---
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Report')
                    ws = writer.book['Report']
                    
                    center_align = Alignment(horizontal='center', vertical='center')
                    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                         top=Side(style='thin'), bottom=Side(style='thin'))
                    colors = ["A2D2FF", "FFD6A5", "CAFFBF", "FDFFB6", "FFADAD", "BDB2FF", "9BF6FF"]
                    
                    last_row = len(df) + 1
                    last_col = len(df.columns)

                    # 1. Align and Borders
                    for r in range(1, last_row + 1):
                        for c in range(1, last_col + 1):
                            cell = ws.cell(row=r, column=c)
                            cell.alignment = center_align
                            cell.border = thin_border

                    # 2. Logic for Merging Location (Column 1)
                    current_loc, start_row_loc = None, 2
                    for row_num in range(2, last_row + 2):
                        row_loc = ws.cell(row=row_num, column=1).value 
                        if row_loc != current_loc or row_num == last_row + 1:
                            if current_loc is not None:
                                end_row_loc = row_num - 1
                                if end_row_loc > start_row_loc:
                                    ws.merge_cells(start_row=start_row_loc, start_column=1, end_row=end_row_loc, end_column=1)
                            start_row_loc, current_loc = row_num, row_loc

                    # 3. Logic for Merging Property (Col 2) and Coloring
                    current_prop, start_row_prop, color_idx = None, 2, 0
                    for row_num in range(2, last_row + 2):
                        row_prop = ws.cell(row=row_num, column=2).value 
                        if row_prop != current_prop or row_num == last_row + 1:
                            if current_prop is not None:
                                end_row_prop = row_num - 1
                                fill = PatternFill(start_color=colors[color_idx % len(colors)], fill_type="solid")
                                for r_fill in range(start_row_prop, end_row_prop + 1):
                                    for c_fill in range(1, last_col + 1):
                                        ws.cell(row=r_fill, column=c_fill).fill = fill
                                if end_row_prop > start_row_prop:
                                    ws.merge_cells(start_row=start_row_prop, start_column=2, end_row=end_row_prop, end_column=2)
                                color_idx += 1
                            start_row_prop, current_prop = row_num, row_prop

                    for col in ws.columns:
                        ws.column_dimensions[col[0].column_letter].width = 20

                file_content = output.getvalue()
                st.success("Professional styling applied!")
                st.dataframe(df.head())

                # --- SIDEBAR EMAIL ---
                st.sidebar.header("📧 Email Report")
                recipient = st.sidebar.text_input("Recipient Name", placeholder="firstname.lastname")
                if st.sidebar.button("Send to Email") and recipient:
                    full_email = f"{recipient.strip().lower()}@beyondwalls.com"
                    if send_email(full_email, file_content, "Professional_Package_Report.xlsx"):
                        st.sidebar.success(f"Sent to {full_email}")


    except Exception as e:
        st.error(f"Error: {e}")
